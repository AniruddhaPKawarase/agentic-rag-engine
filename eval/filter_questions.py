"""Filter the 600-question xlsx down to a topic-focused 50-question set.

Picks questions that are most likely to exercise the v3.1 generation chain
(drawings, floors, equipment, levels, schedules, trades) — the exact topics
the client asked us to optimize for.

Usage:
    python eval/filter_questions.py \\
        --input "<path-to-test_questions_for_all_projects.xlsx>" \\
        --output eval_results/questions_50_focused.xlsx \\
        --count 50

Output is an xlsx in the same single-column shape the eval runner expects.
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from collections import Counter, defaultdict
from typing import Dict, List, Tuple

import openpyxl


# ---------------------------------------------------------------------------
# Topic buckets — keywords matched case-insensitively against question text.
# Each question is scored against each bucket; we pick a balanced mix.
# ---------------------------------------------------------------------------

TOPIC_BUCKETS: Dict[str, List[str]] = {
    "drawings": [
        r"\bdrawing\b",
        r"\bsheet\b",
        r"\bplan\b",
        r"\bdetail\b",
        r"\bsection\b",
        r"\belevation\b",
        r"\briser\b",
        r"\bdiagram\b",
        # Drawing-name patterns: A-101, M101, S-200, E604, P-301
        r"\b[ASMEPC]-?\d{2,3}[a-z]?\b",
    ],
    "floors_levels": [
        r"\bfloor\b",
        r"\blevel\b",
        r"\bstor[ie]y\b",
        r"\bfloor[\s-]to[\s-]floor\b",
        r"\bceiling\b",
        r"\bbasement\b",
        r"\bground\s+floor\b",
        r"\bsecond\s+floor\b",
        r"\bthird\s+floor\b",
        r"\bpenthouse\b",
        r"\broof\b",
        r"\bfoundation\b",
        r"\btypical\b",
    ],
    "equipment": [
        r"\bequipment\b",
        r"\bAHU\b",
        r"\bDOAS\b",
        r"\bRTU\b",
        r"\bVAV\b",
        r"\bchiller\b",
        r"\bboiler\b",
        r"\bpump\b",
        r"\bvalve\b",
        r"\bfan\b",
        r"\bdamper\b",
        r"\bpanel\b",
        r"\bpanelboard\b",
        r"\btransformer\b",
        r"\bswitchboard\b",
        r"\bswitchgear\b",
        r"\bgenerator\b",
        r"\bUPS\b",
        r"\bfixture\b",
        r"\bsensor\b",
        r"\bactuator\b",
        r"\bcompressor\b",
    ],
    "schedules_lists": [
        r"\bschedule\b",
        r"\blist\s+all\b",
        r"\bhow\s+many\b",
        r"\bcount\b",
        r"\btotal\s+number\b",
        r"\bnumber\s+of\b",
        r"\bquantit",
        r"\bidentify\s+all\b",
        r"\benumera",
    ],
    "trades_systems": [
        r"\bmechanical\b",
        r"\belectrical\b",
        r"\bplumbing\b",
        r"\bstructural\b",
        r"\barchitectural\b",
        r"\bcivil\b",
        r"\bfire\s+protection\b",
        r"\bfire\s+alarm\b",
        r"\bHVAC\b",
        r"\blife\s+safety\b",
        r"\bsprinkler\b",
        r"\blighting\b",
        r"\bpower\b",
        r"\bdomestic\s+water\b",
        r"\bwaste\b",
        r"\bvent\b",
    ],
    "spatial_and_quantitative": [
        r"\bdimension",
        r"\bsquare\s+f[oo]+t",
        r"\bsq\.?\s*ft\b",
        r"\barea\b",
        r"\bcapacity\b",
        r"\brating\b",
        r"\bload\b",
        r"\bairflow\b",
        r"\bCFM\b",
        r"\bGPM\b",
        r"\bvoltage\b",
        r"\bamperage\b",
        r"\bclear[\s-]height\b",
    ],
}


# Anti-patterns: questions we want to AVOID (too spec-doc heavy, less drawing-focused)
EXCLUDE_PATTERNS = [
    r"\bDivision\s+0[01]\b",     # admin / general conditions
    r"\bSection\s+01\d",
    r"\ballowance",
    r"\bunit\s+price",
    r"\bsubmittal",
    r"\bwarrant[yi]",
    r"\bcontract\b",
    r"\bclause\b",
    r"\bbid\b",
    r"\baddendum",
]


def score_question(q: str) -> Tuple[int, Dict[str, int]]:
    """Return (total_score, per_bucket_score) for a question.

    Rewards questions that hit drawing/floor/equipment/level/schedule keywords;
    penalizes spec-doc-only patterns we want to avoid.
    """
    q_lower = q.lower()
    bucket_hits: Dict[str, int] = {}
    total = 0

    for bucket, patterns in TOPIC_BUCKETS.items():
        hits = 0
        for pat in patterns:
            if re.search(pat, q, re.IGNORECASE):
                hits += 1
        if hits:
            bucket_hits[bucket] = hits
            total += hits

    for pat in EXCLUDE_PATTERNS:
        if re.search(pat, q, re.IGNORECASE):
            total -= 3  # heavy penalty

    return total, bucket_hits


def load_questions(input_path: str) -> List[str]:
    """Read column A from the xlsx, dropping empty cells."""
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.active
    out: List[str] = []
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        v = row[0]
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        # Skip obvious header rows
        if s.lower() in ("question", "questions", "task", "tasks"):
            continue
        out.append(s)
    return out


def pick_balanced_subset(
    scored: List[Tuple[int, Dict[str, int], str]],
    count: int,
) -> List[Tuple[int, Dict[str, int], str]]:
    """Pick `count` questions balanced across topic buckets.

    Strategy: round-robin across buckets, picking the highest-scoring
    not-yet-picked question that primarily hits each bucket. This avoids
    a final 50 dominated by one topic.
    """
    target_per_bucket = max(1, count // len(TOPIC_BUCKETS))

    # Group by primary bucket (the one with highest hits in that question)
    by_bucket: Dict[str, List[Tuple[int, Dict[str, int], str]]] = defaultdict(list)
    for total, hits, q in scored:
        if total <= 0:
            continue
        primary = max(hits.items(), key=lambda kv: kv[1])[0]
        by_bucket[primary].append((total, hits, q))

    # Sort each bucket by score descending
    for bucket in by_bucket:
        by_bucket[bucket].sort(key=lambda x: -x[0])

    picked: List[Tuple[int, Dict[str, int], str]] = []
    seen: set = set()

    # Round-robin until we hit `count`
    while len(picked) < count:
        added_this_round = False
        for bucket in TOPIC_BUCKETS.keys():
            if len(picked) >= count:
                break
            for entry in by_bucket.get(bucket, []):
                key = entry[2].lower().strip()
                if key in seen:
                    continue
                picked.append(entry)
                seen.add(key)
                added_this_round = True
                break
        if not added_this_round:
            break  # no more candidates anywhere

    return picked[:count]


def write_subset(picked: List[Tuple[int, Dict[str, int], str]], output_path: str) -> None:
    """Write picked questions back to a single-column xlsx (no header)."""
    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "questions"
    # Single column, no header (matches the eval harness's loader behaviour)
    for i, (_, _, q) in enumerate(picked, start=1):
        ws.cell(row=i, column=1, value=q)
    wb.save(output_path)


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--input", required=True, help="Source xlsx (600 questions)")
    p.add_argument("--output", required=True, help="Filtered xlsx (subset)")
    p.add_argument("--count", type=int, default=50, help="How many to pick (default 50)")
    p.add_argument("--show-scores", action="store_true", help="Print scoring breakdown")
    args = p.parse_args()

    if not os.path.exists(args.input):
        print(f"ERROR: input not found: {args.input}", file=sys.stderr)
        return 2

    questions = load_questions(args.input)
    print(f"Loaded {len(questions)} questions from {args.input}")

    scored: List[Tuple[int, Dict[str, int], str]] = []
    for q in questions:
        score, hits = score_question(q)
        scored.append((score, hits, q))

    positive = [s for s in scored if s[0] > 0]
    print(f"  {len(positive)} questions hit at least one target bucket")
    print(f"  {len(scored) - len(positive)} questions excluded (no hits or penalised)")

    picked = pick_balanced_subset(scored, args.count)

    # Per-bucket breakdown for the chosen set
    bucket_counts: Counter = Counter()
    for _, hits, _ in picked:
        primary = max(hits.items(), key=lambda kv: kv[1])[0] if hits else "none"
        bucket_counts[primary] += 1

    print(f"\nPicked {len(picked)} questions, distribution by primary topic:")
    for bucket, n in sorted(bucket_counts.items(), key=lambda kv: -kv[1]):
        print(f"  {bucket:<26}  {n:>3}")

    if args.show_scores:
        print("\nSelected questions (top 10 shown):")
        for i, (score, hits, q) in enumerate(picked[:10], 1):
            buckets_str = ", ".join(f"{b}={n}" for b, n in hits.items())
            print(f"  {i:>2}. score={score:>2} [{buckets_str}]")
            print(f"       {q[:140]}")

    write_subset(picked, args.output)
    print(f"\nWrote {len(picked)} questions to {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
