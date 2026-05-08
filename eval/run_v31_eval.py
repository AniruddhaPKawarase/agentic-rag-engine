"""v3.1 generation chain offline evaluation harness — Phase P5 (Path A).

Pure local Python imports. Loads the v3.1 chain modules directly,
toggles env flags between v3.0 baseline (flags off) and v3.1 variant
(flags on), runs questions through the chain end-to-end with real LLM
calls, and writes CSV + JSON reports.

Modes:
    single_turn — 600-question xlsx through both baseline + variant
    multi_turn  — 50 hand-crafted reference-resolution pairs through variant
    model_ab    — single_turn variant repeated with 3 model sets
    all         — single_turn + multi_turn + model_ab + EVAL_REPORT.md

Usage:
    python eval/run_v31_eval.py --mode single_turn --questions <xlsx> \\
        --project-id 7325 --set-id 4987 --sample 30 \\
        --output-dir eval_results/single_turn --model-set haiku
"""

from __future__ import annotations

import argparse
import asyncio
import contextlib
import csv
import json
import logging
import os
import re
import statistics
import sys
import time
import traceback
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

# ---------------------------------------------------------------------------
# Worktree path injection — the harness must work from any CWD.
# ---------------------------------------------------------------------------
HERE = os.path.dirname(os.path.abspath(__file__))
WORKTREE = os.path.dirname(HERE)
if WORKTREE not in sys.path:
    sys.path.insert(0, WORKTREE)

# Load .env BEFORE any module that reads os.environ.get("ANTHROPIC_API_KEY")
# / "OPENAI_API_KEY" / "MONGODB_URI" is imported. Otherwise the LLM clients
# see missing keys and silently fall back to OpenAI even when the eval is
# meant to test Claude.
#
# Lookup order (first match for each variable wins via override=False):
#   1. v3.1 worktree .env  (sandboxed config for this version)
#   2. unified-rag-agent root .env  (the live-tree config; user puts the
#      Anthropic key here so it's also picked up when v3.1 deploys)
try:
    from dotenv import load_dotenv  # type: ignore
    _CANDIDATE_ENVS = [
        os.path.join(WORKTREE, ".env"),
        os.path.join(os.path.dirname(os.path.dirname(WORKTREE)), ".env"),  # repo root
    ]
    for _env_path in _CANDIDATE_ENVS:
        if os.path.exists(_env_path):
            load_dotenv(_env_path, override=False)
except ImportError:
    pass

logger = logging.getLogger("agentic_rag.eval.runner")


# ---------------------------------------------------------------------------
# Hand-crafted multi-turn reference-resolution pairs (50 total).
# ---------------------------------------------------------------------------

# Each pair = (turn1_query, turn2_query). Project-7325-realistic
# (electrical / mechanical / plumbing trades).

MULTI_TURN_PAIRS: List[Tuple[str, str]] = [
    # ── Group A — Pronoun resolution (15) ────────────────────────────────
    ("List all electrical drawings.", "What about the second one?"),
    ("Show me drawing M-501.", "And the schedules on it?"),
    ("What's the foundation type?", "Is it the same on level 2?"),
    ("List all panel schedules.", "What's on the third one?"),
    ("Show me the AHU on level 3.", "What's its CFM?"),
    ("List all transformers.", "Where is the largest one located?"),
    ("Show drawing E-101.", "What loads are shown on it?"),
    ("List all fire alarm devices.", "How many of those are on level 1?"),
    ("Show me the main switchgear.", "What feeders come off it?"),
    ("List all plumbing fixtures in the lobby.", "Are they ADA-compliant?"),
    ("Show drawing P-201.", "What's the pipe size on it?"),
    ("List all VAVs.", "Which one serves conference room 2?"),
    ("Show the chiller specs.", "What's its tonnage?"),
    ("List all roof drains.", "What size are they?"),
    ("Show the main electrical room.", "What panels are inside it?"),
    # ── Group B — Trade follow-up (15) ───────────────────────────────────
    ("List all electrical equipment.", "Now do the same for plumbing."),
    ("Show me the AHU schedules.", "And the chiller schedule?"),
    ("List CSI Division 26 specs.", "What about Division 23?"),
    ("Show all electrical panels.", "Same for distribution panels?"),
    ("List mechanical equipment on level 1.", "Same for level 2."),
    ("Show me all VAVs on level 3.", "Now do the FCUs."),
    ("List all junction boxes.", "And the pull boxes?"),
    ("Show plumbing risers.", "Same for sanitary risers?"),
    ("List all light fixtures.", "Now the emergency lights."),
    ("Show HVAC controls on level 2.", "Same for level 3."),
    ("List all motors.", "Now the motor starters."),
    ("Show fire protection devices.", "And the fire alarm devices?"),
    ("List electrical panel schedules.", "Same for the lighting schedules."),
    ("Show all dampers.", "Now the smoke dampers specifically."),
    ("List all conduits over 2 inches.", "Same for under 2 inches."),
    # ── Group C — Implicit count / reference (10) ────────────────────────
    ("How many DOAS units are there?", "How many of those are on level 3?"),
    ("Show all fire dampers.", "Group them by drawing."),
    ("How many AHUs are in the project?", "What's their average CFM?"),
    ("List all transformers.", "How many are 75kVA or larger?"),
    ("Show all plumbing fixtures.", "How many water closets?"),
    ("How many VAVs are there?", "Which level has the most?"),
    ("List all pumps.", "How many are duty/standby pairs?"),
    ("Show all FCUs.", "Average them by tonnage."),
    ("How many panels are 480V?", "Which one has the highest amperage?"),
    ("List all motorized dampers.", "Group them by fan system."),
    # ── Group D — Continuation (10) ──────────────────────────────────────
    ("What is the floor-to-floor height?", "And the ceiling height?"),
    ("List CSI Division 23 specs.", "What about Division 26?"),
    ("Show the main lobby finishes.", "And the corridor finishes?"),
    ("What's the roof slope?", "And the parapet height?"),
    ("Show foundation details.", "What about the slab thickness?"),
    ("What's the typical wall assembly?", "And the floor assembly?"),
    ("Show the curtainwall details.", "And the storefront details?"),
    ("List exterior door types.", "What about interior doors?"),
    ("Show the stair details.", "And the elevator shaft details?"),
    ("What's the structural steel grade?", "And the rebar grade?"),
]


# ---------------------------------------------------------------------------
# Model sets for model_ab mode.
# ---------------------------------------------------------------------------

MODEL_SETS: Dict[str, Dict[str, str]] = {
    "haiku": {
        "SYNTHESIZER_MODEL": "claude-haiku-4-5",
        "SYNTHESIZER_MODEL_FALLBACK": "gpt-4o-mini",
        "STYLIST_MODEL": "claude-haiku-4-5",
        "STYLIST_MODEL_FALLBACK": "gpt-4o-mini",
    },
    "sonnet": {
        "SYNTHESIZER_MODEL": "claude-sonnet-4-6",
        "SYNTHESIZER_MODEL_FALLBACK": "gpt-4o",
        "STYLIST_MODEL": "claude-sonnet-4-6",
        "STYLIST_MODEL_FALLBACK": "gpt-4o",
    },
    "gpt4o-mini": {
        "SYNTHESIZER_MODEL": "gpt-4o-mini",
        "SYNTHESIZER_MODEL_FALLBACK": "claude-haiku-4-5",
        "STYLIST_MODEL": "gpt-4o-mini",
        "STYLIST_MODEL_FALLBACK": "claude-haiku-4-5",
    },
}


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------


@dataclass
class ChainRunMetrics:
    """Metrics collected from one chain invocation (baseline or variant)."""

    question_index: int
    question: str
    project_id: int
    set_id: Optional[int]
    answer: str = ""
    answer_length: int = 0
    source_count: int = 0
    citation_count: int = 0
    latency_ms: int = 0
    cost_usd: float = 0.0
    model_used: str = ""
    contextualized_query: str = ""
    synthesizer_used: bool = False
    stylist_used: bool = False
    memory_context_used: bool = False
    error: str = ""

    def as_csv_row(self) -> Dict[str, Any]:
        return asdict(self)


# ---------------------------------------------------------------------------
# Env flag helpers
# ---------------------------------------------------------------------------


@contextlib.contextmanager
def env_flags(**flags: Any):
    """Temporarily set env vars; restore on exit. Values are str-coerced."""
    old = {k: os.environ.get(k) for k in flags}
    for k, v in flags.items():
        if v is None:
            os.environ.pop(k, None)
        elif isinstance(v, bool):
            os.environ[k] = "true" if v else "false"
        else:
            os.environ[k] = str(v)
    try:
        yield
    finally:
        for k, v in old.items():
            if v is None:
                os.environ.pop(k, None)
            else:
                os.environ[k] = v


def baseline_flags() -> Dict[str, str]:
    """v3.0 baseline — every v3.1 flag OFF, master switch OFF."""
    return {
        "V31_CHAIN_ENABLED": "false",
        "MEMORY_RECALL_ENABLED": "false",
        "QUERY_REWRITER_ENABLED": "false",
        "ANSWER_SYNTHESIZER_ENABLED": "false",
        "STYLE_REWRITER_ENABLED": "false",
        "CACHE_REEXPRESSION_ENABLED": "false",
        "MEMORY_WRITER_VECTOR_ENABLED": "false",
    }


def variant_flags(*, with_memory: bool = False) -> Dict[str, str]:
    """v3.1 variant — synthesizer + stylist always on. Memory only for multi-turn."""
    return {
        "V31_CHAIN_ENABLED": "true",
        "MEMORY_RECALL_ENABLED": "true" if with_memory else "false",
        "QUERY_REWRITER_ENABLED": "true" if with_memory else "false",
        "ANSWER_SYNTHESIZER_ENABLED": "true",
        "STYLE_REWRITER_ENABLED": "true",
        "CACHE_REEXPRESSION_ENABLED": "false",
        "MEMORY_WRITER_VECTOR_ENABLED": "true" if with_memory else "false",
    }


# ---------------------------------------------------------------------------
# Question loading
# ---------------------------------------------------------------------------


def load_questions(xlsx_path: str) -> List[str]:
    """Load questions from an xlsx with either a header row or a bare column.

    Looks for columns named ``question`` or ``Task`` (case-insensitive),
    falling back to the first non-empty column. Drops blanks.
    """
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover — env issue
        raise RuntimeError(f"openpyxl required to load xlsx: {exc}") from exc

    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    first_row = rows[0]
    has_header = False
    target_col = 0
    if first_row:
        for idx, cell in enumerate(first_row):
            if cell is None:
                continue
            cell_str = str(cell).strip().lower()
            if cell_str in ("question", "task", "questions"):
                has_header = True
                target_col = idx
                break

    data_rows = rows[1:] if has_header else rows
    questions: List[str] = []
    for row in data_rows:
        if not row:
            continue
        # Pick the first non-empty cell in the row if header didn't lock a col
        if has_header:
            cell = row[target_col] if target_col < len(row) else None
        else:
            cell = next((c for c in row if c is not None and str(c).strip()), None)
        if cell is None:
            continue
        text = str(cell).strip()
        if text:
            questions.append(text)
    return questions


# ---------------------------------------------------------------------------
# Citation / metrics helpers
# ---------------------------------------------------------------------------

# Match drawing-style citations like [E-101], [M-501.2], [P-201, p.3], etc.
_CITATION_RE = re.compile(r"\[[A-Z]{1,3}-?\d[\w\-.,\s/]{0,40}\]")


def count_citations(text: str) -> int:
    """Count bracketed citation markers in answer text."""
    if not text:
        return 0
    return len(_CITATION_RE.findall(text))


def percent(part: float, whole: float) -> float:
    if not whole:
        return 0.0
    return round(100.0 * part / whole, 2)


def safe_median(values: List[float]) -> float:
    return round(statistics.median(values), 2) if values else 0.0


def safe_p95(values: List[float]) -> float:
    if not values:
        return 0.0
    sorted_v = sorted(values)
    idx = max(0, int(len(sorted_v) * 0.95) - 1)
    return round(sorted_v[idx], 2)


# ---------------------------------------------------------------------------
# Core chain invocation
# ---------------------------------------------------------------------------


def _import_chain():
    """Late import so env flags are picked up correctly per call."""
    from gateway.generation_chain import run_generation_chain  # noqa: WPS433

    return run_generation_chain


def run_chain_once(
    *,
    question: str,
    question_index: int,
    project_id: int,
    set_id: Optional[int],
    session_id: Optional[str] = None,
    flags: Optional[Dict[str, str]] = None,
    extra_env: Optional[Dict[str, str]] = None,
) -> ChainRunMetrics:
    """Invoke the chain end-to-end with the given env flags. Never raises.

    Errors are captured into the returned ChainRunMetrics.error field.
    """
    flags = flags or {}
    extra_env = extra_env or {}
    merged: Dict[str, Any] = {**flags, **extra_env}
    metrics = ChainRunMetrics(
        question_index=question_index,
        question=question,
        project_id=project_id,
        set_id=set_id,
    )

    started = time.monotonic()
    try:
        with env_flags(**merged):
            run_generation_chain = _import_chain()
            coro = run_generation_chain(
                user_query=question,
                session_id=session_id,
                project_id=project_id,
                set_id=set_id,
                scope=None,
                cached_result=None,
                conversation_history=None,
                stream=False,
            )
            result: Dict[str, Any] = asyncio.run(coro)
    except Exception as exc:  # noqa: BLE001 — log + record, never propagate
        elapsed = int((time.monotonic() - started) * 1000)
        metrics.latency_ms = elapsed
        metrics.error = f"{type(exc).__name__}: {exc}"
        logger.warning(
            "run_chain_once failed q_idx=%d err=%s\n%s",
            question_index,
            metrics.error,
            traceback.format_exc(limit=2),
        )
        return metrics

    elapsed = int((time.monotonic() - started) * 1000)
    answer = result.get("answer", "") or ""
    sources = result.get("sources", []) or []
    debug = result.get("debug_info", {}) or {}

    metrics.answer = answer
    metrics.answer_length = len(answer)
    metrics.source_count = len(sources)
    metrics.citation_count = count_citations(answer)
    metrics.latency_ms = elapsed
    metrics.cost_usd = float(debug.get("agentic_cost_usd", 0.0) or 0.0)
    metrics.model_used = result.get("model_used", "") or ""
    metrics.contextualized_query = result.get("contextualized_query", question) or question
    metrics.synthesizer_used = bool(result.get("synthesizer_used", False))
    metrics.stylist_used = bool(result.get("stylist_used", False))
    metrics.memory_context_used = bool(result.get("memory_context_used", False))
    if result.get("error"):
        metrics.error = str(result.get("error"))
    return metrics


# ---------------------------------------------------------------------------
# CSV writers
# ---------------------------------------------------------------------------

_CSV_FIELDS = [
    "question_index",
    "question",
    "project_id",
    "set_id",
    "answer",
    "answer_length",
    "source_count",
    "citation_count",
    "latency_ms",
    "cost_usd",
    "model_used",
    "contextualized_query",
    "synthesizer_used",
    "stylist_used",
    "memory_context_used",
    "error",
]


def write_metrics_csv(rows: List[ChainRunMetrics], out_path: Path) -> None:
    """Write a list of ChainRunMetrics to CSV with a fixed schema."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=_CSV_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row.as_csv_row())


def read_metrics_csv(path: Path) -> List[ChainRunMetrics]:
    """Re-hydrate ChainRunMetrics rows from a previous run."""
    rows: List[ChainRunMetrics] = []
    with open(path, "r", encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        for raw in reader:
            rows.append(
                ChainRunMetrics(
                    question_index=int(raw.get("question_index") or 0),
                    question=raw.get("question") or "",
                    project_id=int(raw.get("project_id") or 0),
                    set_id=int(raw["set_id"]) if raw.get("set_id") else None,
                    answer=raw.get("answer") or "",
                    answer_length=int(raw.get("answer_length") or 0),
                    source_count=int(raw.get("source_count") or 0),
                    citation_count=int(raw.get("citation_count") or 0),
                    latency_ms=int(raw.get("latency_ms") or 0),
                    cost_usd=float(raw.get("cost_usd") or 0.0),
                    model_used=raw.get("model_used") or "",
                    contextualized_query=raw.get("contextualized_query") or "",
                    synthesizer_used=raw.get("synthesizer_used") == "True",
                    stylist_used=raw.get("stylist_used") == "True",
                    memory_context_used=raw.get("memory_context_used") == "True",
                    error=raw.get("error") or "",
                )
            )
    return rows


# ---------------------------------------------------------------------------
# Aggregation
# ---------------------------------------------------------------------------


def summarize_pair(
    baseline: List[ChainRunMetrics],
    variant: List[ChainRunMetrics],
) -> Dict[str, Any]:
    """Compute the single_turn_summary.json contents."""
    by_idx_b = {m.question_index: m for m in baseline}
    by_idx_v = {m.question_index: m for m in variant}
    common = sorted(set(by_idx_b) & set(by_idx_v))

    b_lens = [by_idx_b[i].answer_length for i in common if not by_idx_b[i].error]
    v_lens = [by_idx_v[i].answer_length for i in common if not by_idx_v[i].error]
    b_lat = [by_idx_b[i].latency_ms for i in common if not by_idx_b[i].error]
    v_lat = [by_idx_v[i].latency_ms for i in common if not by_idx_v[i].error]
    b_cost = [by_idx_b[i].cost_usd for i in common if not by_idx_b[i].error]
    v_cost = [by_idx_v[i].cost_usd for i in common if not by_idx_v[i].error]

    citations_kept = 0
    citations_compared = 0
    variant_shorter = 0
    for i in common:
        bi, vi = by_idx_b[i], by_idx_v[i]
        if bi.error or vi.error:
            continue
        if bi.citation_count > 0:
            citations_compared += 1
            if vi.citation_count >= bi.citation_count:
                citations_kept += 1
        if vi.answer_length < bi.answer_length:
            variant_shorter += 1

    total_pairs = sum(1 for i in common if not by_idx_b[i].error and not by_idx_v[i].error)

    return {
        "total_questions": len(common),
        "total_pairs_clean": total_pairs,
        "baseline_errors": sum(1 for m in baseline if m.error),
        "variant_errors": sum(1 for m in variant if m.error),
        "length": {
            "baseline_median": safe_median(b_lens),
            "variant_median": safe_median(v_lens),
            "baseline_p95": safe_p95(b_lens),
            "variant_p95": safe_p95(v_lens),
            "variant_shorter_pct": percent(variant_shorter, total_pairs),
        },
        "latency_ms": {
            "baseline_median": safe_median(b_lat),
            "variant_median": safe_median(v_lat),
            "baseline_p95": safe_p95(b_lat),
            "variant_p95": safe_p95(v_lat),
            "delta_median": safe_median(v_lat) - safe_median(b_lat),
        },
        "cost_usd": {
            "baseline_total": round(sum(b_cost), 4),
            "variant_total": round(sum(v_cost), 4),
            "delta_total": round(sum(v_cost) - sum(b_cost), 4),
        },
        "citations": {
            "compared": citations_compared,
            "kept": citations_kept,
            "kept_pct": percent(citations_kept, citations_compared),
        },
        "synthesizer_fired_pct": percent(
            sum(1 for m in variant if m.synthesizer_used), len(variant)
        ),
        "stylist_fired_pct": percent(
            sum(1 for m in variant if m.stylist_used), len(variant)
        ),
    }


# ---------------------------------------------------------------------------
# Mode runners
# ---------------------------------------------------------------------------


def _run_parallel(
    questions: List[str],
    fn,
    max_workers: int,
    label: str,
) -> List[ChainRunMetrics]:
    """Run ``fn(idx, question)`` over a list with a thread pool."""
    results: List[Optional[ChainRunMetrics]] = [None] * len(questions)
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futs = {
            pool.submit(fn, idx, q): idx
            for idx, q in enumerate(questions)
        }
        completed = 0
        total = len(futs)
        for fut in as_completed(futs):
            idx = futs[fut]
            try:
                results[idx] = fut.result()
            except Exception as exc:  # noqa: BLE001
                logger.error("[%s] worker idx=%d crashed: %s", label, idx, exc)
                results[idx] = ChainRunMetrics(
                    question_index=idx,
                    question=questions[idx],
                    project_id=-1,
                    set_id=None,
                    error=f"worker_crash: {exc}",
                )
            completed += 1
            if completed % 5 == 0 or completed == total:
                logger.info("[%s] progress %d/%d", label, completed, total)
    return [r for r in results if r is not None]


def run_single_turn_mode(
    *,
    questions: List[str],
    project_id: int,
    set_id: Optional[int],
    output_dir: Path,
    max_workers: int,
    model_set: str,
    baseline_only: bool,
    variant_only: bool,
) -> Dict[str, Any]:
    """Drive the single_turn comparison: baseline vs variant for each Q."""
    output_dir.mkdir(parents=True, exist_ok=True)
    extra_env = MODEL_SETS.get(model_set, MODEL_SETS["haiku"])

    baseline_rows: List[ChainRunMetrics] = []
    variant_rows: List[ChainRunMetrics] = []

    if not variant_only:
        logger.info("Running baseline (v3.0, all flags off) — %d questions", len(questions))

        def _baseline_fn(idx: int, q: str) -> ChainRunMetrics:
            return run_chain_once(
                question=q,
                question_index=idx,
                project_id=project_id,
                set_id=set_id,
                session_id=None,
                flags=baseline_flags(),
                extra_env=None,
            )

        baseline_rows = _run_parallel(questions, _baseline_fn, max_workers, "baseline")
        write_metrics_csv(baseline_rows, output_dir / "single_turn_baseline.csv")
    else:
        baseline_csv = output_dir / "single_turn_baseline.csv"
        if baseline_csv.exists():
            baseline_rows = read_metrics_csv(baseline_csv)
            logger.info("Reusing existing baseline csv (%d rows)", len(baseline_rows))
        else:
            logger.warning("--variant-only set but no baseline csv at %s", baseline_csv)

    if not baseline_only:
        logger.info("Running variant (v3.1, flags on) — %d questions", len(questions))

        def _variant_fn(idx: int, q: str) -> ChainRunMetrics:
            return run_chain_once(
                question=q,
                question_index=idx,
                project_id=project_id,
                set_id=set_id,
                session_id=None,
                flags=variant_flags(with_memory=False),
                extra_env=extra_env,
            )

        variant_rows = _run_parallel(questions, _variant_fn, max_workers, "variant")
        write_metrics_csv(variant_rows, output_dir / "single_turn_variant.csv")

    summary: Dict[str, Any] = {}
    if baseline_rows and variant_rows:
        summary = summarize_pair(baseline_rows, variant_rows)
        summary["model_set"] = model_set
        summary["mode"] = "single_turn"
        with open(output_dir / "single_turn_summary.json", "w", encoding="utf-8") as fh:
            json.dump(summary, fh, indent=2)
    return summary


def run_multi_turn_mode(
    *,
    project_id: int,
    set_id: Optional[int],
    output_dir: Path,
    pairs: List[Tuple[str, str]],
) -> Dict[str, Any]:
    """Drive the multi_turn coherence test (variant + memory on)."""
    output_dir.mkdir(parents=True, exist_ok=True)

    from eval.llm_judge import score_multi_turn_coherence  # noqa: WPS433

    flags = variant_flags(with_memory=True)
    out_rows: List[Dict[str, Any]] = []
    scores: List[int] = []

    for i, (turn1_q, turn2_q) in enumerate(pairs):
        session_id = f"eval-mt-{i}-{uuid.uuid4().hex[:6]}"
        logger.info("[multi_turn] pair %d/%d session=%s", i + 1, len(pairs), session_id)

        t1 = run_chain_once(
            question=turn1_q,
            question_index=2 * i,
            project_id=project_id,
            set_id=set_id,
            session_id=session_id,
            flags=flags,
        )
        # Small delay to let memory writer flush before turn 2 reads.
        time.sleep(1.0)
        t2 = run_chain_once(
            question=turn2_q,
            question_index=2 * i + 1,
            project_id=project_id,
            set_id=set_id,
            session_id=session_id,
            flags=flags,
        )

        if t1.error or t2.error:
            score = -1
        else:
            score = score_multi_turn_coherence(turn1_q, t1.answer, turn2_q, t2.answer)
        if score >= 1:
            scores.append(score)

        out_rows.append({
            "pair_index": i,
            "session_id": session_id,
            "turn1_query": turn1_q,
            "turn1_answer": t1.answer,
            "turn1_error": t1.error,
            "turn2_query": turn2_q,
            "turn2_answer": t2.answer,
            "turn2_contextualized": t2.contextualized_query,
            "turn2_memory_used": t2.memory_context_used,
            "turn2_error": t2.error,
            "judge_score": score,
            "turn1_latency_ms": t1.latency_ms,
            "turn2_latency_ms": t2.latency_ms,
        })

    csv_path = output_dir / "multi_turn_results.csv"
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as fh:
        if out_rows:
            writer = csv.DictWriter(fh, fieldnames=list(out_rows[0].keys()))
            writer.writeheader()
            writer.writerows(out_rows)

    summary = {
        "mode": "multi_turn",
        "total_pairs": len(pairs),
        "scored": len(scores),
        "median_score": safe_median([float(s) for s in scores]),
        "pct_score_ge_4": percent(sum(1 for s in scores if s >= 4), len(scores)),
        "pct_score_ge_3": percent(sum(1 for s in scores if s >= 3), len(scores)),
        "memory_used_pct": percent(
            sum(1 for r in out_rows if r["turn2_memory_used"]), len(out_rows)
        ),
    }
    with open(output_dir / "multi_turn_summary.json", "w", encoding="utf-8") as fh:
        json.dump(summary, fh, indent=2)
    return summary


def run_model_ab_mode(
    *,
    questions: List[str],
    project_id: int,
    set_id: Optional[int],
    output_dir: Path,
    max_workers: int,
) -> Dict[str, Any]:
    """Run variant 3 times with 3 model sets; LLM-judge tone."""
    output_dir.mkdir(parents=True, exist_ok=True)
    from eval.llm_judge import score_human_tone  # noqa: WPS433

    by_set: Dict[str, List[ChainRunMetrics]] = {}
    for set_name in ("haiku", "sonnet", "gpt4o-mini"):
        logger.info("[model_ab] running variant with model set %s", set_name)
        extra_env = MODEL_SETS[set_name]

        def _fn(idx: int, q: str, _env=extra_env) -> ChainRunMetrics:
            return run_chain_once(
                question=q,
                question_index=idx,
                project_id=project_id,
                set_id=set_id,
                session_id=None,
                flags=variant_flags(with_memory=False),
                extra_env=_env,
            )

        rows = _run_parallel(questions, _fn, max_workers, f"model_ab/{set_name}")
        write_metrics_csv(rows, output_dir / f"model_ab_{set_name}.csv")
        by_set[set_name] = rows

    # LLM-judge tone per row per set
    judge_rows: List[Dict[str, Any]] = []
    for idx, q in enumerate(questions):
        row = {"question_index": idx, "question": q}
        for set_name in ("haiku", "sonnet", "gpt4o-mini"):
            metrics = by_set[set_name][idx] if idx < len(by_set[set_name]) else None
            if metrics and not metrics.error and metrics.answer:
                row[f"{set_name}_score"] = score_human_tone(q, metrics.answer)
                row[f"{set_name}_length"] = metrics.answer_length
                row[f"{set_name}_latency_ms"] = metrics.latency_ms
            else:
                row[f"{set_name}_score"] = -1
                row[f"{set_name}_length"] = 0
                row[f"{set_name}_latency_ms"] = 0
        judge_rows.append(row)

    csv_path = output_dir / "model_ab_results.csv"
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as fh:
        if judge_rows:
            writer = csv.DictWriter(fh, fieldnames=list(judge_rows[0].keys()))
            writer.writeheader()
            writer.writerows(judge_rows)

    def _avg_score(name: str) -> float:
        vals = [r[f"{name}_score"] for r in judge_rows if r[f"{name}_score"] >= 1]
        return round(sum(vals) / len(vals), 2) if vals else 0.0

    summary = {
        "mode": "model_ab",
        "model_sets": list(MODEL_SETS.keys()),
        "averages": {
            name: {
                "tone_score_avg": _avg_score(name),
                "length_median": safe_median(
                    [r[f"{name}_length"] for r in judge_rows]
                ),
                "latency_median_ms": safe_median(
                    [r[f"{name}_latency_ms"] for r in judge_rows]
                ),
                "total_cost_usd": round(
                    sum(m.cost_usd for m in by_set[name]), 4
                ),
            }
            for name in MODEL_SETS
        },
    }
    with open(output_dir / "model_ab_summary.json", "w", encoding="utf-8") as fh:
        json.dump(summary, fh, indent=2)
    return summary


def write_master_report(
    *,
    output_dir: Path,
    single_turn: Optional[Dict[str, Any]],
    multi_turn: Optional[Dict[str, Any]],
    model_ab: Optional[Dict[str, Any]],
) -> None:
    """Compose EVAL_REPORT.md summarising all sub-modes."""
    lines: List[str] = ["# v3.1 Eval Report\n"]

    if single_turn:
        lines.append("## Single-turn (v3.1 vs v3.0)\n")
        lines.append(f"- Questions compared: {single_turn.get('total_pairs_clean', 0)}")
        ln = single_turn.get("length", {})
        lat = single_turn.get("latency_ms", {})
        cost = single_turn.get("cost_usd", {})
        cit = single_turn.get("citations", {})
        lines.append(
            f"- Length median (b/v): {ln.get('baseline_median')} -> {ln.get('variant_median')} "
            f"({ln.get('variant_shorter_pct')}% shorter)"
        )
        lines.append(
            f"- Latency median (b/v): {lat.get('baseline_median')}ms -> {lat.get('variant_median')}ms "
            f"(delta {lat.get('delta_median')}ms)"
        )
        lines.append(
            f"- Cost total (b/v): ${cost.get('baseline_total')} -> ${cost.get('variant_total')} "
            f"(delta ${cost.get('delta_total')})"
        )
        lines.append(
            f"- Citations kept: {cit.get('kept')}/{cit.get('compared')} ({cit.get('kept_pct')}%)"
        )
        lines.append(
            f"- Synthesizer fired: {single_turn.get('synthesizer_fired_pct')}%; "
            f"Stylist fired: {single_turn.get('stylist_fired_pct')}%\n"
        )

    if multi_turn:
        lines.append("## Multi-turn coherence\n")
        lines.append(f"- Pairs: {multi_turn.get('total_pairs')}; scored: {multi_turn.get('scored')}")
        lines.append(f"- Median judge score: {multi_turn.get('median_score')}/5")
        lines.append(f"- Score >=4: {multi_turn.get('pct_score_ge_4')}%; >=3: {multi_turn.get('pct_score_ge_3')}%")
        lines.append(f"- Memory context used in turn-2: {multi_turn.get('memory_used_pct')}%\n")

    if model_ab:
        lines.append("## Model A/B (variant only)\n")
        avgs = model_ab.get("averages", {})
        best = max(avgs.items(), key=lambda kv: kv[1].get("tone_score_avg", 0)) if avgs else None
        for name, agg in avgs.items():
            lines.append(
                f"- **{name}**: tone {agg.get('tone_score_avg')}/5, "
                f"len_med {agg.get('length_median')}, "
                f"lat_med {agg.get('latency_median_ms')}ms, "
                f"cost ${agg.get('total_cost_usd')}"
            )
        if best:
            lines.append(f"\n**Recommended model set:** `{best[0]}`\n")

    target_lines: List[str] = ["## Target check\n"]
    if single_turn:
        ln = single_turn.get("length", {})
        cit = single_turn.get("citations", {})
        target_lines.append(
            f"- Variant shorter than baseline >= 60%: "
            f"{'YES' if ln.get('variant_shorter_pct', 0) >= 60 else 'NO'} "
            f"({ln.get('variant_shorter_pct')}%)"
        )
        target_lines.append(
            f"- Citation preservation >= 95%: "
            f"{'YES' if cit.get('kept_pct', 0) >= 95 else 'NO'} "
            f"({cit.get('kept_pct')}%)"
        )
    if multi_turn:
        target_lines.append(
            f"- Multi-turn median >= 4: "
            f"{'YES' if multi_turn.get('median_score', 0) >= 4 else 'NO'} "
            f"({multi_turn.get('median_score')})"
        )
    lines.extend(target_lines)

    out_path = output_dir / "EVAL_REPORT.md"
    out_path.write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# Cost / runtime banner
# ---------------------------------------------------------------------------


def print_banner(*, mode: str, sample: int, output_dir: Path) -> None:
    """Print upfront cost + runtime banner; sleep 5s so user can abort."""
    if mode == "single_turn":
        runs = sample * 2
        cost = sample * 0.01  # ~$0.01 per chain call (haiku + retrieval)
        runtime_min = max(5, int(runs / 6))  # 3 workers * ~17-22s
    elif mode == "multi_turn":
        runs = len(MULTI_TURN_PAIRS) * 2
        cost = len(MULTI_TURN_PAIRS) * 0.025  # 2 chain calls + judge
        runtime_min = max(8, int(runs / 3))
    elif mode == "model_ab":
        runs = sample * 3
        cost = sample * 0.04  # 3x variant + judge per question
        runtime_min = max(10, int(runs / 6))
    else:  # all
        runs = sample * 2 + len(MULTI_TURN_PAIRS) * 2 + sample * 3
        cost = sample * 0.05 + len(MULTI_TURN_PAIRS) * 0.025
        runtime_min = max(20, int(runs / 6))

    print("=== v3.1 EVAL HARNESS ===")
    print(f"Mode: {mode}")
    print(f"Sample size: {sample} questions")
    print(f"Estimated chain runs: {runs}")
    print(f"Estimated LLM cost: ${cost:.2f} (range ${cost*0.7:.2f}-${cost*1.7:.2f})")
    print(f"Estimated runtime: ~{runtime_min} minutes")
    print(f"Output: {output_dir}")
    print("Press Ctrl+C in 5 seconds to abort...")
    try:
        time.sleep(5)
    except KeyboardInterrupt:
        print("Aborted.")
        sys.exit(1)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="run_v31_eval",
        description="v3.1 generation chain offline evaluation harness (Phase P5).",
    )
    p.add_argument(
        "--mode",
        required=True,
        choices=["single_turn", "multi_turn", "model_ab", "all"],
        help="Eval mode to run.",
    )
    p.add_argument(
        "--questions",
        type=str,
        default=None,
        help="Path to xlsx with the question set (single_turn / model_ab / all).",
    )
    p.add_argument("--project-id", type=int, required=True, help="Project ID.")
    p.add_argument("--set-id", type=int, default=None, help="Drawing-set ID.")
    p.add_argument(
        "--sample",
        type=int,
        default=0,
        help="Sample size (0 = use full set).",
    )
    p.add_argument("--output-dir", type=str, required=True, help="Output directory.")
    p.add_argument(
        "--model-set",
        type=str,
        default="haiku",
        choices=list(MODEL_SETS.keys()),
        help="Which model set powers Agents 4+5 in single_turn mode.",
    )
    p.add_argument("--max-workers", type=int, default=3, help="Thread-pool size.")
    p.add_argument(
        "--baseline-only",
        action="store_true",
        help="Run only the v3.0 baseline (skip variant).",
    )
    p.add_argument(
        "--variant-only",
        action="store_true",
        help="Run only the v3.1 variant; reuse existing baseline csv.",
    )
    p.add_argument(
        "--no-banner",
        action="store_true",
        help="Skip the cost / runtime banner + 5s confirmation pause.",
    )
    return p


def _select_questions(
    *,
    args: argparse.Namespace,
    require: bool = True,
) -> List[str]:
    """Load + sample the question set."""
    if not args.questions:
        if require:
            raise SystemExit("--questions is required for this mode.")
        return []
    path = args.questions
    if not os.path.exists(path):
        raise SystemExit(f"Questions file not found: {path}")
    questions = load_questions(path)
    if args.sample and args.sample > 0:
        questions = questions[: args.sample]
    return questions


def main(argv: Optional[List[str]] = None) -> int:
    logging.basicConfig(
        level=os.environ.get("EVAL_LOG_LEVEL", "INFO"),
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    output_dir = Path(args.output_dir).resolve()

    sample_size = args.sample or 30
    if not args.no_banner:
        print_banner(mode=args.mode, sample=sample_size, output_dir=output_dir)

    single_summary: Optional[Dict[str, Any]] = None
    multi_summary: Optional[Dict[str, Any]] = None
    model_ab_summary: Optional[Dict[str, Any]] = None

    if args.mode in ("single_turn", "all"):
        questions = _select_questions(args=args)
        single_summary = run_single_turn_mode(
            questions=questions,
            project_id=args.project_id,
            set_id=args.set_id,
            output_dir=output_dir if args.mode == "single_turn" else output_dir / "single_turn",
            max_workers=args.max_workers,
            model_set=args.model_set,
            baseline_only=args.baseline_only,
            variant_only=args.variant_only,
        )

    if args.mode in ("multi_turn", "all"):
        multi_summary = run_multi_turn_mode(
            project_id=args.project_id,
            set_id=args.set_id,
            output_dir=output_dir if args.mode == "multi_turn" else output_dir / "multi_turn",
            pairs=MULTI_TURN_PAIRS,
        )

    if args.mode in ("model_ab", "all"):
        questions = _select_questions(args=args)
        model_ab_summary = run_model_ab_mode(
            questions=questions,
            project_id=args.project_id,
            set_id=args.set_id,
            output_dir=output_dir if args.mode == "model_ab" else output_dir / "model_ab",
            max_workers=args.max_workers,
        )

    if args.mode == "all":
        write_master_report(
            output_dir=output_dir,
            single_turn=single_summary,
            multi_turn=multi_summary,
            model_ab=model_ab_summary,
        )
        logger.info("EVAL_REPORT.md written to %s", output_dir / "EVAL_REPORT.md")

    print("DONE.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
